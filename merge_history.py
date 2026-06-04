"""
merge_history.py — Fusiona el reporte nuevo con el historial acumulado.

Flujo:
  1. Lee 'latest.xlsx'  (descargado por download.js, últimos 7 días)
  2. Lee 'INFORME EXCESOS DE VELOCIDAD.xlsx' si existe (historial)
  3. Agrega filas nuevas (dedup por Alias + Fecha de Inicio)
  4. Guarda como 'INFORME EXCESOS DE VELOCIDAD.xlsx'
"""

import openpyxl
import os
import shutil
from datetime import datetime

LATEST   = 'latest.xlsx'
HIST     = 'INFORME EXCESOS DE VELOCIDAD.xlsx'
MAX_DAYS = 60   # ventana rodante: siempre los últimos 60 días

def get_header_row(ws, max_search=10):
    """Devuelve (row_idx, col_map) para la hoja."""
    for r in range(1, max_search + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(str(v).strip() == 'Alias' for v in row_vals if v):
            col = {str(v).strip(): i+1 for i, v in enumerate(row_vals) if v}
            return r, col
    return None, {}

def extract_rows(ws, header_row, col):
    """Extrae lista de dicts con todos los valores de cada fila de datos."""
    rows = []
    headers = list(col.keys())
    for r in range(header_row + 1, ws.max_row + 1):
        alias = ws.cell(r, col.get('Alias', 1)).value
        if not alias:
            continue
        row_data = {h: ws.cell(r, col[h]).value for h in headers}
        rows.append(row_data)
    return rows

def row_key(row_data):
    """Clave única: Alias + Fecha de Inicio exacta."""
    return (str(row_data.get('Alias', '')).strip(),
            str(row_data.get('Fecha de Inicio', '')).strip())

def parse_fecha(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ('%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(str(val), fmt)
        except ValueError:
            pass
    return None

def main():
    # ── Determinar archivos de entrada ────────────────────────────────────────
    import glob as glob_mod
    bulk_files = sorted(glob_mod.glob('bulk_*.xlsx'))
    source_files = bulk_files if bulk_files else ([LATEST] if os.path.exists(LATEST) else [])

    if not source_files:
        print(f'ERROR: no se encontraron archivos bulk_*.xlsx ni {LATEST}')
        exit(1)

    # ── Leer todos los archivos nuevos ────────────────────────────────────────
    new_rows = []
    for src in source_files:
        try:
            wb_src = openpyxl.load_workbook(src)
            if 'Detalle 1' not in wb_src.sheetnames:
                print(f'  [SKIP] {src}: sin hoja Detalle 1')
                continue
            ws_src = wb_src['Detalle 1']
            hr_src, col_src = get_header_row(ws_src)
            if not hr_src:
                print(f'  [SKIP] {src}: sin encabezado')
                continue
            rows_src = extract_rows(ws_src, hr_src, col_src)
            new_rows.extend(rows_src)
            print(f'  {src}: {len(rows_src)} filas')
        except Exception as e:
            print(f'  [WARN] {src}: {e}')
    print(f'Total filas nuevas: {len(new_rows)}')

    # ── Leer historial existente (si hay) ─────────────────────────────────────
    existing_rows = []
    existing_keys = set()
    if os.path.exists(HIST):
        wb_old = openpyxl.load_workbook(HIST)
        if 'Detalle 1' in wb_old.sheetnames:
            ws_old = wb_old['Detalle 1']
            hr_old, col_old = get_header_row(ws_old)
            if hr_old:
                existing_rows = extract_rows(ws_old, hr_old, col_old)
                existing_keys = {row_key(r) for r in existing_rows}
                print(f'Historial existente: {len(existing_rows)} filas')

    # ── Fusionar: agregar solo filas nuevas ───────────────────────────────────
    added = 0
    for row in new_rows:
        k = row_key(row)
        if k not in existing_keys:
            existing_rows.append(row)
            existing_keys.add(k)
            added += 1
    print(f'Filas nuevas agregadas: {added}')

    # ── Ordenar por Fecha de Inicio ───────────────────────────────────────────
    existing_rows.sort(key=lambda r: parse_fecha(r.get('Fecha de Inicio')) or datetime.min)

    # ── Limitar a MAX_DAYS días ───────────────────────────────────────────────
    if existing_rows:
        last_fecha = parse_fecha(existing_rows[-1].get('Fecha de Inicio'))
        if last_fecha:
            cutoff = last_fecha.replace(hour=0, minute=0, second=0, microsecond=0)
            from datetime import timedelta
            cutoff = cutoff - timedelta(days=MAX_DAYS)
            before = len(existing_rows)
            existing_rows = [r for r in existing_rows
                             if (parse_fecha(r.get('Fecha de Inicio')) or datetime.min) >= cutoff]
            print(f'Recorte historial: {before} → {len(existing_rows)} filas (límite {MAX_DAYS} días)')

    # ── Construir el Excel final copiando estructura del nuevo reporte ─────────
    shutil.copy(LATEST, HIST)
    wb_out = openpyxl.load_workbook(HIST)
    ws_out = wb_out['Detalle 1']
    hr_out, col_out = get_header_row(ws_out)

    # Borrar filas de datos existentes
    ws_out.delete_rows(hr_out + 1, ws_out.max_row - hr_out)

    # Escribir filas fusionadas
    headers_out = list(col_out.keys())
    for i, row_data in enumerate(existing_rows):
        r_idx = hr_out + 1 + i
        for h in headers_out:
            ws_out.cell(r_idx, col_out[h]).value = row_data.get(h)

    # Actualizar celda de período en Resumen 1
    if 'Resumen 1' in wb_out.sheetnames and existing_rows:
        ws_r = wb_out['Resumen 1']
        fecha_ini = parse_fecha(existing_rows[0].get('Fecha de Inicio'))
        fecha_fin = parse_fecha(existing_rows[-1].get('Fecha de Inicio'))
        for r in range(1, min(10, ws_r.max_row + 1)):
            v = ws_r.cell(r, 1).value
            if v and 'Desde' in str(v):
                if fecha_ini and fecha_fin:
                    ws_r.cell(r, 1).value = (
                        f"Desde : {fecha_ini.strftime('%Y/%m/%d')} 00:00:00 "
                        f"Hasta: {fecha_fin.strftime('%Y/%m/%d')} 23:59:59"
                    )
                break

    wb_out.save(HIST)
    print(f'✅ {HIST} actualizado con {len(existing_rows)} filas en total')

if __name__ == '__main__':
    main()
